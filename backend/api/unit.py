from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import joinedload
from typing import List, Union
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from backend.db import DbSessionDep
from backend.tables import (
    NppUnitTable,
    PlacementTable,
    CouponComplectTable,
    ContainerSysTable,
    CouponLoadTable,
    CouponExtractTable,
)

unit_router = APIRouter()


class PlacementWithHistory:
    """
    Helper class to track the history of loads and extracts for a placement.
    """

    def __init__(self, placement: PlacementTable):
        self.placement = placement
        self.history: List[Union[CouponLoadTable, CouponExtractTable]] = []

    def load(self, load: CouponLoadTable):
        """Add a load event to the history."""
        self.history.append(load)

    def extract(self, extract: CouponExtractTable):
        """Add an extract event to the history."""
        self.history.append(extract)


class ContainerSysWithHistory:
    """
    Helper class to track the history of loads and extracts for a container system.
    """

    def __init__(self, container_sys: ContainerSysTable):
        self.container_sys = container_sys
        self.history: List[Union[CouponLoadTable, CouponExtractTable]] = []

    def load(self, load: CouponLoadTable):
        """Add a load event to the history."""
        self.history.append(load)

    def extract(self, extract: CouponExtractTable):
        """Add an extract event to the history."""
        self.history.append(extract)


placements_coords = {
    1: {
        1: (5377, 2432),
        2: (5313, 2194),
        3: (5120, 1775),
        4: (4853, 1399),
        5: (4678, 1225),
    },
    2: {
        1: (3696, 658),
        2: (3459, 595),
        3: (2999, 552),
        4: (2539, 595),
        5: (2303, 658)
    },
    3: {
        1: (1320, 1225),
        2: (1146, 1399),
        3: (879, 1775),
        4: (687, 2194),
        5: (623, 2432),
    },
    4: {
        1: (623, 3566),
        2: (687, 3805),
        3: (879, 4223),
        4: (1146, 4599),
        5: (1320, 4773),
    },
    5: {
        1: (2303, 5341),
        2: (2539, 5404),
        3: (2999, 5448),
        4: (3459, 5404),
        5: (3696, 5341),
    },
    6: {
        1: (4678, 4773),
        2: (4853, 4599),
        3: (5120, 4223),
        4: (5313, 3805),
        5: (5377, 3566),
    },
}  # fmt: skip


placement_text_coords = {
    1: {
        1: (5712, 2376),
        2: (5642, 2114),
        3: (5431, 1653),
        4: (5137, 1240),
        5: (4946, 1048),
    },
    2: {
        1: (3810, 425),
        2: (3470, 355),
        3: (3015, 317),
        4: (2540, 355),
        5: (2210, 425)
    },
    3: {
        1: (1103, 1048),
        2: (912, 1240),
        3: (618, 1653),
        4: (407, 2114),
        5: (337, 2376),
    },
    4: {
        1: (337, 3623),
        2: (407, 3885),
        3: (618, 4346),
        4: (912, 4759),
        5: (1103, 4951),
    },
    5: {
        1: (2210, 5574),
        2: (2540, 5646),
        3: (3015, 5693),
        4: (3470, 5645),
        5: (3810, 5574),
    },
    6: {
        1: (4946, 4951),
        2: (5137, 4759),
        3: (5431, 4346),
        4: (5642, 3885),
        5: (5712, 3623),
    },
}  # fmt: skip


@unit_router.get("/unit/{name_eng}", operation_id="get_unit")
def unit_detail(name_eng: str, db: DbSessionDep):
    """
    Get specific unit by name_eng with complete placement and complects data.
    """

    # Get unit by name_eng
    unit = db.query(NppUnitTable).filter(NppUnitTable.name_eng == name_eng).first()

    if unit is None:
        raise HTTPException(status_code=404, detail="Unit not found")

    # Get placements for this unit, ordered by sector and sector_num
    placements_data = (
        db.query(PlacementTable)
        .filter(PlacementTable.unit_id == unit.unit_id)
        .order_by(PlacementTable.sector, PlacementTable.sector_num)
        .all()
    )

    # Initialize placement history tracking
    placements = {}
    for placement in placements_data:
        ph = PlacementWithHistory(placement)
        placements[placement.placement_id] = ph

    # Container system history tracking
    container_sys_history = {}

    # Fetch all complects for the unit with their related container systems
    complects_data = (
        db.query(CouponComplectTable)
        .filter(CouponComplectTable.unit_id == unit.unit_id)
        .options(joinedload(CouponComplectTable.container_systems))
        .order_by(CouponComplectTable.is_additional.asc(), CouponComplectTable.name)
        .all()
    )

    complects = {}
    for complect in complects_data:
        complect_systems = []
        # Sort container systems by name
        sorted_container_systems = sorted(
            complect.container_systems, key=lambda cs: cs.name
        )
        for container_sys in sorted_container_systems:
            csh = ContainerSysWithHistory(container_sys)
            complect_systems.append(csh)
            container_sys_history[container_sys.container_sys_id] = csh
        complects[complect.name] = {
            "systems": complect_systems,
            "is_additional": complect.is_additional,
        }
    loads = (
        db.query(CouponLoadTable)
        .join(
            PlacementTable,
            CouponLoadTable.irrad_placement_id == PlacementTable.placement_id,
        )
        .join(
            ContainerSysTable,
            CouponLoadTable.irrad_container_sys_id
            == ContainerSysTable.container_sys_id,
        )
        .join(
            CouponComplectTable,
            ContainerSysTable.coupon_complect_id
            == CouponComplectTable.coupon_complect_id,
        )
        .filter(PlacementTable.unit_id == unit.unit_id)
        .options(
            joinedload(CouponLoadTable.irrad_container_sys).joinedload(
                ContainerSysTable.coupon_complect
            ),
            joinedload(CouponLoadTable.irrad_placement),
            joinedload(CouponLoadTable.irrad_container_sys).joinedload(
                ContainerSysTable.coupon_extracts
            ),
        )
        .order_by(
            CouponLoadTable.load_date, ContainerSysTable.name, CouponComplectTable.name
        )
        .all()
    )

    # Process each load and its corresponding extract
    for load in loads:
        # Add load to placement history
        placement_id = load.irrad_placement.placement_id
        ph = placements[placement_id]
        ph.load(load)

        # Add load to container system history
        container_sys_id = load.irrad_container_sys.container_sys_id
        cs = container_sys_history[container_sys_id]
        cs.load(load)

        # Get extract for this container system (0 or 1 extract per container system)
        extract = None
        for ext in load.irrad_container_sys.coupon_extracts:
            extract = ext
            break  # Take first (should be only one)

        if extract is not None:
            ph.extract(extract)
            cs.extract(extract)

    # Build response data structure matching Django view
    return {
        "unit": {
            "unit_id": unit.unit_id,
            "plant_id": unit.plant_id,
            "num": int(unit.num),
            "name": unit.name,
            "name_eng": unit.name_eng,
            "design": unit.design,
            "stage": unit.stage,
            "power": float(unit.power) if unit.power else None,
            "start_date": unit.start_date.isoformat() if unit.start_date else None,
        },
        "placements": [
            {
                "history": [
                    (
                        {
                            "type": "load",
                            "container_sys_name": event.irrad_container_sys.name,
                            "date": event.load_date if event.load_date else None,
                        }
                        if isinstance(event, CouponLoadTable)
                        else {
                            "type": "extract",
                            "container_sys_name": event.irrad_container_sys.name,
                            "date": event.extract_date if event.extract_date else None,
                        }
                    )
                    for event in ph.history
                ],
                "placement": {
                    "name": ph.placement.name,
                    "sector": int(ph.placement.sector),
                    "sector_num": int(ph.placement.sector_num),
                },
            }
            for ph in placements.values()
        ],
        "placements_coords": placements_coords,
        "placements_text_coords": placement_text_coords,
        "complects": {
            complect_name: {
                "is_additional": complect_data["is_additional"],
                "systems": [
                    {
                        "container_sys_name": csh.container_sys.name,
                        "history": [
                            (
                                {
                                    "type": "load",
                                    "load_id": event.cpn_load_id,
                                    "date": event.load_date
                                    if event.load_date
                                    else None,
                                    "placement_name": event.irrad_placement.name,
                                }
                                if isinstance(event, CouponLoadTable)
                                else {
                                    "type": "extract",
                                    "extract_id": event.cpn_extract_id,
                                    "date": event.extract_date
                                    if event.extract_date
                                    else None,
                                }
                            )
                            for event in csh.history
                        ],
                    }
                    for csh in complect_data["systems"]
                ],
            }
            for complect_name, complect_data in complects.items()
        },
    }


@unit_router.get("/unit/{name_eng}/export", operation_id="export_unit_data")
def export_unit_data(name_eng: str, db: DbSessionDep):
    """
    Export unit data to Excel file.
    """
    # Get unit data using the same logic as unit_detail
    unit = db.query(NppUnitTable).filter(NppUnitTable.name_eng == name_eng).first()

    if unit is None:
        raise HTTPException(status_code=404, detail="Unit not found")

    # Get all the data (reusing logic from unit_detail)
    unit_data = unit_detail(name_eng, db)  # Create Excel workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    default_sheet = wb.active
    if default_sheet:
        wb.remove(default_sheet)

    # Create Unit Info sheet
    ws_unit = wb.create_sheet("Інформація про блок")

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )

    # Unit information
    unit_info = unit_data["unit"]
    ws_unit.append(["Параметр", "Значення"])

    # Style header
    for cell in ws_unit[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Add unit data
    ws_unit.append(["Номер блоку", unit_info.get("num", "-")])
    ws_unit.append(["Найменування блоку", unit_info.get("name", "-")])
    ws_unit.append(["Найменування блоку (англ.)", unit_info.get("name_eng", "-")])
    ws_unit.append(["Проект", unit_info.get("design", "-")])
    ws_unit.append(["Черга", unit_info.get("stage", "-")])
    ws_unit.append(
        [
            "Встановлена потужність, МВ",
            round(unit_info["power"]) if unit_info.get("power") else "-",
        ]
    )
    ws_unit.append(
        [
            "Дата початку експлуатації",
            datetime.fromisoformat(unit_info["start_date"]).strftime("%d.%m.%Y")
            if unit_info.get("start_date")
            else "-",
        ]
    )  # Auto-adjust column widths
    for column in ws_unit.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_unit.column_dimensions[column_letter].width = adjusted_width

    # Create Placements sheet
    ws_placements = wb.create_sheet("Історія місць")

    # Placements header
    ws_placements.append(["Місце", "Збірка", "Завантажено", "Вивантажено"])

    # Style header
    for cell in ws_placements[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Process placements data
    for placement_data in unit_data["placements"]:
        placement_name = placement_data["placement"]["name"]

        # Track load/extract pairs
        history = placement_data["history"]
        history_periods = []
        last_load_event = None

        for event in history:
            if event["type"] == "load":
                if last_load_event is not None:
                    raise ValueError("Load event into occupied placement!")
                last_load_event = {
                    "container_sys_name": event["container_sys_name"],
                    "load_date": event["date"],
                }
            else:  # extract
                if last_load_event is None:
                    raise ValueError("Extract event from empty placement!")
                history_periods.append(
                    {
                        "container_sys_name": last_load_event["container_sys_name"],
                        "load_date": last_load_event["load_date"],
                        "extract_date": event["date"],
                    }
                )
                last_load_event = None

        # Add current load if exists
        if last_load_event is not None:
            history_periods.append(
                {
                    "container_sys_name": last_load_event["container_sys_name"],
                    "load_date": last_load_event["load_date"],
                    "extract_date": None,
                }
            )

        # Sort by load date (assuming string dates in sortable format like YYYY-MM-DD)
        history_periods.sort(key=lambda x: x["load_date"] if x["load_date"] else "")

        # Add rows to sheet
        for period in history_periods:
            load_date = period["load_date"] if period["load_date"] else ""
            extract_date = (
                period["extract_date"] if period["extract_date"] else "опромінюється"
            )

            ws_placements.append(
                [placement_name, period["container_sys_name"], load_date, extract_date]
            )  # Auto-adjust column widths
    for column in ws_placements.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_placements.column_dimensions[column_letter].width = adjusted_width

    # Create Complects sheets
    for complect_name, complect_data in unit_data["complects"].items():
        # Create sheet for each complect
        ws_complect = wb.create_sheet(f"Комплект {complect_name}")

        # Header
        ws_complect.append(["Збірка", "Завантажено", "Місце", "Вивантажено"])

        # Style header
        for cell in ws_complect[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Collect all periods from all container systems in this complect
        all_periods = []

        for container_sys in complect_data["systems"]:
            container_sys_name = container_sys["container_sys_name"]
            history = container_sys["history"]

            # Process history in pairs (load, extract)
            i = 0
            while i < len(history):
                if history[i]["type"] == "load":
                    load_event = history[i]
                    extract_event = (
                        history[i + 1]
                        if i + 1 < len(history) and history[i + 1]["type"] == "extract"
                        else None
                    )

                    load_date = load_event["date"] if load_event["date"] else ""
                    extract_date = (
                        extract_event["date"] if extract_event else "опромінюється"
                    )

                    all_periods.append(
                        {
                            "container_sys_name": container_sys_name,
                            "load_date": load_event["date"],
                            "load_date_formatted": load_date,
                            "placement_name": load_event["placement_name"],
                            "extract_date_formatted": extract_date,
                        }
                    )

                    i += 2 if extract_event else 1
                else:
                    i += 1

        # Sort by load date and placement
        all_periods.sort(key=lambda x: (x["load_date"], x["placement_name"]))

        # Add rows
        for period in all_periods:
            ws_complect.append(
                [
                    period["container_sys_name"],
                    period["load_date_formatted"],
                    period["placement_name"],
                    period["extract_date_formatted"],
                ]
            )  # Auto-adjust column widths
        for column in ws_complect.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws_complect.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"unit_{name_eng}_{timestamp}.xlsx"

    # Return as streaming response
    return StreamingResponse(
        BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
